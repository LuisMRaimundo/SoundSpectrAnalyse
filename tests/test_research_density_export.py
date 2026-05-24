"""Tests for ``tools/export_research_density_workbook.py``."""

from __future__ import annotations

import hashlib
import subprocess
import sys
import zipfile
from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[1]


def test_research_workbook_has_no_xl_table_parts_and_loads(tmp_path: Path) -> None:
    """Formal Excel Table parts (xl/tables/table*.xml) trigger Excel repair; export must not create them."""
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "compiled_density_metrics_research.xlsx"
    _write_minimal_compiled_workbook(src)
    proc = _run_export(src, dst)
    assert proc.returncode == 0, proc.stderr + proc.stdout
    with zipfile.ZipFile(dst) as z:
        table_parts = [n for n in z.namelist() if n.startswith("xl/tables/")]
        assert table_parts == []
    wb = load_workbook(dst)
    assert "Dashboard" in wb.sheetnames
    assert "Spectral_Density_Metrics" in wb.sheetnames
    sdm = wb["Spectral_Density_Metrics"]
    assert sdm.auto_filter.ref is not None
    assert str(sdm.auto_filter.ref).startswith("A1:")


def test_sanitize_dataframe_columns_blank_and_duplicates() -> None:
    from tools.export_research_density_workbook import _sanitize_dataframe_columns

    df = pd.DataFrame([[1, 2, 3]], columns=["", "Note", "Note"])
    out = _sanitize_dataframe_columns(df)
    assert list(out.columns) == ["column_1", "Note", "Note_2"]


def _write_minimal_compiled_workbook(path: Path, *, sparse: bool = False) -> None:
    """Minimal ``compiled_density_metrics``-style workbook for export tests."""
    dm = pd.DataFrame(
        {
            "Note": ["A4", "D3"],
            "density_metric_raw": [1.0, 1.0],
            "density_metric_normalized": [1.0, 1.0],
            "weighted_harmonic_density_contribution": [0.2, 0.8],
            "weighted_inharmonic_density_contribution": [0.3, 0.15],
            "weighted_subbass_density_contribution": [0.5, 0.05],
            "component_harmonic_energy_ratio": [0.2, 0.8],
            "component_inharmonic_energy_ratio": [0.3, 0.15],
            "component_subbass_energy_ratio": [0.5, 0.05],
            "Harmonic Partials sum": [1.0, 10.0],
            "Inharmonic Partials sum": [2.0, 20.0],
            "Sub-bass sum": [3.0, 30.0],
            "Total sum": [6.0, 60.0],
            "source_file_name": ["Clarinet_A4_pp.wav", "Bassoon_D3_mf.wav"],
            "weight_function": ["linear", "linear"],
            "density_weighted_sum": [1.5, 2.5],
            "Combined Density Metric": [4.0, 8.0],
            "density_log_weighted": [0.1, 0.2],
            "effective_partial_density": [0.4, 0.5],
            "spectral_entropy": [0.7, 0.8],
        }
    )
    if sparse:
        dm = pd.DataFrame({"Note": ["C4"], "density_metric_raw": [0.5]})

    am = pd.DataFrame(
        [
            ("pipeline_contract_version", "test-contract"),
            ("ANALYSIS_SCHEMA_VERSION", "99"),
            ("weight_function", "linear"),
            ("window_type", "blackmanharris"),
            ("n_fft", 4096),
            ("hop_length", 1024),
            ("zero_padding", 2),
            ("harmonic_tolerance", 5.0),
            ("frequency_min_hz", 20.0),
            ("frequency_max_hz", 20000.0),
            ("magnitude_min_db", -90.0),
            ("density_summation_mode", "his_weighted"),
            ("harmonic_density_weight", 1.0),
            ("inharmonic_density_weight", 0.5),
            ("subbass_density_weight", 0.25),
            ("density_salience_threshold_db", -45.0),
            ("density_frequency_ceiling_hz", 5000.0),
        ],
        columns=["Parameter", "Value"],
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        dm.to_excel(writer, sheet_name="Density_Metrics", index=False)
        am.to_excel(writer, sheet_name="Analysis_Metadata", index=False)


def _run_export(inp: Path, out: Path, extra: list[str] | None = None) -> subprocess.CompletedProcess[str]:
    cmd = [
        sys.executable,
        str(REPO / "tools" / "export_research_density_workbook.py"),
        str(inp),
        "--output",
        str(out),
        "--overwrite",
    ]
    if extra:
        cmd.extend(extra)
    return subprocess.run(cmd, capture_output=True, text=True, check=False)


def test_export_creates_research_workbook(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "compiled_density_metrics_research.xlsx"
    _write_minimal_compiled_workbook(src)
    proc = _run_export(src, dst)
    assert proc.returncode == 0, proc.stderr + proc.stdout
    assert dst.is_file()
    xl = pd.ExcelFile(dst, engine="openpyxl")
    expected = {
        "README",
        "Dashboard",
        "Spectral_Density_Metrics",
        "Legacy_Compatibility",
        "Component_Balance",
        "Validation_Summary",
        "Charts_Data",
        "Analysis_Settings_By_Note",
        "Metadata",
    }
    assert set(xl.sheet_names) == expected


def test_spectral_density_metrics_columns(tmp_path: Path) -> None:
    src = tmp_path / "in.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    assert _run_export(src, dst).returncode == 0
    df = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    for col in (
        "Note",
        "MIDI",
        "density_metric_raw",
        "density_weighted_sum",
        "Total sum",
        "effective_partial_density",
        "spectral_entropy",
    ):
        assert col in df.columns
    assert "density_weighted_sum_cdm_mean" not in df.columns


def test_research_workbook_column_highlights(tmp_path: Path) -> None:
    src = tmp_path / "in.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    assert _run_export(src, dst, extra=["--include-legacy-cdm-mean"]).returncode == 0
    from tools.export_research_density_workbook import RESEARCH_FILL_DENSITY_WEIGHTED_SUM, RESEARCH_FILL_DWS_CDM_MEAN

    wb = load_workbook(dst)
    ws = wb["Spectral_Density_Metrics"]
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    def _fill_rgb_hex(cell) -> str:
        col = cell.fill.fgColor
        if col is None:
            return ""
        rgb = getattr(col, "rgb", None) or getattr(getattr(cell.fill, "start_color", None), "rgb", None)
        if rgb is None:
            return ""
        s = str(rgb).upper().replace("FF", "", 1) if str(rgb).upper().startswith("FF") else str(rgb).upper()
        return s[-6:] if len(s) >= 6 else s

    def fill_for(name: str) -> str:
        ci = hdr[name]
        return _fill_rgb_hex(ws.cell(2, ci))

    assert fill_for("density_weighted_sum") == "D6E4F0"
    assert "Combined Density Metric" not in hdr

    legacy_ws = wb["Legacy_Compatibility"]
    legacy_hdr = {legacy_ws.cell(1, c).value: c for c in range(1, legacy_ws.max_column + 1)}
    assert "density_weighted_sum_cdm_mean" in legacy_hdr


def test_legacy_cdm_mean_is_opt_in(tmp_path: Path) -> None:
    src = tmp_path / "in.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    assert _run_export(src, dst, extra=["--include-legacy-cdm-mean"]).returncode == 0
    df = pd.read_excel(dst, sheet_name="Legacy_Compatibility", engine="openpyxl")
    assert "density_weighted_sum_cdm_mean" in df.columns
    mean = pd.to_numeric(df["density_weighted_sum_cdm_mean"], errors="coerce")
    assert mean.notna().any()


def test_legacy_compatibility_midi_aligned_by_note_not_row_index(tmp_path: Path) -> None:
    src = tmp_path / "in_midi.xlsx"
    dst = tmp_path / "out_midi.xlsx"
    dm = pd.DataFrame(
        {
            "Note": ["D6", "C#4", "A#3", "D3", "C2", "A2"],
            "density_metric_raw": [1, 2, 3, 4, 5, 6],
            "density_weighted_sum": [1, 2, 3, 4, 5, 6],
            "Combined Density Metric": [10, 20, 30, 40, 50, 60],
            "Weighted Combined Metric": [11, 21, 31, 41, 51, 61],
            "Total Metric": [12, 22, 32, 42, 52, 62],
        }
    )
    am = pd.DataFrame([("ANALYSIS_SCHEMA_VERSION", "99")], columns=["Parameter", "Value"])
    with pd.ExcelWriter(src, engine="openpyxl") as writer:
        dm.to_excel(writer, sheet_name="Density_Metrics", index=False)
        am.to_excel(writer, sheet_name="Analysis_Metadata", index=False)
    assert _run_export(src, dst).returncode == 0
    legacy = pd.read_excel(dst, sheet_name="Legacy_Compatibility", engine="openpyxl")
    got = {
        str(r["Note"]): int(r["MIDI"])
        for _, r in legacy.iterrows()
        if pd.notna(r.get("Note")) and pd.notna(r.get("MIDI"))
    }
    assert got.get("D3") == 50
    assert got.get("A#3") == 58
    assert got.get("C#4") == 61
    assert got.get("D6") == 86
    assert got.get("C2") == 36
    assert got.get("A2") == 45


def test_spectral_density_metrics_note_midi_mapping_stable(tmp_path: Path) -> None:
    src = tmp_path / "in_sdm.xlsx"
    dst = tmp_path / "out_sdm.xlsx"
    _write_minimal_compiled_workbook(src)
    assert _run_export(src, dst).returncode == 0
    sdm = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    from tools.export_research_density_workbook import note_to_midi

    if {"Note", "MIDI"}.issubset(sdm.columns):
        expected = sdm["Note"].map(note_to_midi)
        got = pd.to_numeric(sdm["MIDI"], errors="coerce")
        mask = expected.notna() & got.notna()
        if mask.any():
            assert np.allclose(pd.to_numeric(expected[mask], errors="coerce"), got[mask], atol=1e-9)


def test_component_balance_recomputes(tmp_path: Path) -> None:
    src = tmp_path / "in.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    assert _run_export(src, dst).returncode == 0
    cb = pd.read_excel(dst, sheet_name="Component_Balance", engine="openpyxl")
    wsum = (
        pd.to_numeric(cb["weighted_harmonic_density_contribution"], errors="coerce")
        + pd.to_numeric(cb["weighted_inharmonic_density_contribution"], errors="coerce")
        + pd.to_numeric(cb["weighted_subbass_density_contribution"], errors="coerce")
    )
    assert np.allclose(
        pd.to_numeric(cb["density_metric_raw_recomputed"], errors="coerce"),
        wsum,
        equal_nan=True,
    )
    tsum = (
        pd.to_numeric(cb["harmonic_density_sum"], errors="coerce")
        + pd.to_numeric(cb["inharmonic_density_sum"], errors="coerce")
        + pd.to_numeric(cb["subbass_density_sum"], errors="coerce")
    )
    assert np.allclose(
        pd.to_numeric(cb["total_sum_recomputed"], errors="coerce"),
        tsum,
        equal_nan=True,
    )


def test_source_workbook_not_modified(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    h_before = hashlib.sha256(src.read_bytes()).hexdigest()
    assert _run_export(src, dst).returncode == 0
    h_after = hashlib.sha256(src.read_bytes()).hexdigest()
    assert h_before == h_after


def test_sparse_workbook_no_crash_and_readme_warnings(tmp_path: Path) -> None:
    src = tmp_path / "sparse.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src, sparse=True)
    proc = _run_export(src, dst)
    assert proc.returncode == 0, proc.stderr
    assert "WARNING" in proc.stderr or "Min-max" in proc.stderr or "Component_Balance" in proc.stderr
    readme = pd.read_excel(dst, sheet_name="README", engine="openpyxl", header=None)
    text = "\n".join(str(x) for x in readme.iloc[:, 0].tolist() if pd.notna(x))
    assert "Warnings" in text


def test_charts_data_sorted_by_midi(tmp_path: Path) -> None:
    src = tmp_path / "in.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    assert _run_export(src, dst).returncode == 0
    cd = pd.read_excel(dst, sheet_name="Charts_Data", engine="openpyxl")
    midi = pd.to_numeric(cd["MIDI"], errors="coerce")
    assert midi.is_monotonic_increasing


def test_no_charts_dashboard_has_no_chart_objects(tmp_path: Path) -> None:
    src = tmp_path / "in.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    proc = _run_export(src, dst, extra=["--no-charts"])
    assert proc.returncode == 0
    wb = load_workbook(dst)
    assert "Dashboard" in wb.sheetnames
    dash = wb["Dashboard"]
    assert len(getattr(dash, "_charts", [])) == 0


def test_default_output_path_fails_if_exists(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics.xlsx"
    _write_minimal_compiled_workbook(src)
    default_out = tmp_path / "compiled_density_metrics_research.xlsx"
    default_out.write_text("block")
    proc = subprocess.run(
        [sys.executable, str(REPO / "tools" / "export_research_density_workbook.py"), str(src)],
        capture_output=True,
        text=True,
        check=False,
    )
    assert proc.returncode != 0
    assert "already exists" in proc.stderr.lower() or "already exists" in proc.stdout.lower()


def _write_compiled_with_instrument_dynamic(path: Path) -> None:
    dm = pd.DataFrame(
        {
            "Note": ["A4", "D3"],
            "density_metric_raw": [1.0, 1.0],
            "density_metric_normalized": [1.0, 1.0],
            "weighted_harmonic_density_contribution": [0.2, 0.8],
            "weighted_inharmonic_density_contribution": [0.3, 0.15],
            "weighted_subbass_density_contribution": [0.5, 0.05],
            "component_harmonic_energy_ratio": [0.2, 0.8],
            "component_inharmonic_energy_ratio": [0.3, 0.15],
            "component_subbass_energy_ratio": [0.5, 0.05],
            "Harmonic Partials sum": [1.0, 10.0],
            "Inharmonic Partials sum": [2.0, 20.0],
            "Sub-bass sum": [3.0, 30.0],
            "Total sum": [6.0, 60.0],
            "source_file_name": ["take1.wav", "take2.wav"],
            "Instrument": ["Bassoon", "Bassoon"],
            "Dynamic": ["pp", "pp"],
            "weight_function": ["linear", "linear"],
            "density_weighted_sum": [1.5, 2.5],
            "density_log_weighted": [0.1, 0.2],
            "effective_partial_density": [0.4, 0.5],
            "spectral_entropy": [0.7, 0.8],
        }
    )
    am = pd.DataFrame(
        [
            ("pipeline_contract_version", "test-contract"),
            ("ANALYSIS_SCHEMA_VERSION", "99"),
            ("weight_function", "linear"),
        ],
        columns=["Parameter", "Value"],
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        dm.to_excel(writer, sheet_name="Density_Metrics", index=False)
        am.to_excel(writer, sheet_name="Analysis_Metadata", index=False)


def test_cli_instrument_dynamic_populate_when_absent(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    proc = _run_export(
        src,
        dst,
        extra=["--instrument", "Oboe", "--dynamic", "pp"],
    )
    assert proc.returncode == 0, proc.stderr + proc.stdout
    df = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    assert (df["Instrument"].astype(str) == "Oboe").all()
    assert (df["Dynamic"].astype(str) == "pp").all()
    cb = pd.read_excel(dst, sheet_name="Component_Balance", engine="openpyxl")
    assert (cb["Instrument"].astype(str) == "Oboe").all()
    assert (cb["Dynamic"].astype(str) == "pp").all()
    vs = pd.read_excel(dst, sheet_name="Validation_Summary", engine="openpyxl")
    assert (vs["Instrument"].astype(str) == "Oboe").all()


def test_cli_without_force_does_not_override_existing_instrument_dynamic(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_compiled_with_instrument_dynamic(src)
    proc = _run_export(
        src,
        dst,
        extra=["--instrument", "Flute", "--dynamic", "fff"],
    )
    assert proc.returncode == 0, proc.stderr + proc.stdout
    df = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    assert (df["Instrument"].astype(str) == "Bassoon").all()
    assert (df["Dynamic"].astype(str) == "pp").all()


def test_force_metadata_overrides_existing(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_compiled_with_instrument_dynamic(src)
    proc = _run_export(
        src,
        dst,
        extra=["--instrument", "Clarinet", "--dynamic", "mf", "--force-metadata"],
    )
    assert proc.returncode == 0, proc.stderr + proc.stdout
    df = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    assert (df["Instrument"].astype(str) == "Clarinet").all()
    assert (df["Dynamic"].astype(str) == "mf").all()


def test_inference_from_filename_tokens(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    assert _run_export(src, dst).returncode == 0
    df = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    assert df.loc[df["Note"] == "A4", "Instrument"].iloc[0] == "Clarinet"
    assert df.loc[df["Note"] == "A4", "Dynamic"].iloc[0] == "pp"
    assert df.loc[df["Note"] == "D3", "Instrument"].iloc[0] == "Bassoon"
    assert df.loc[df["Note"] == "D3", "Dynamic"].iloc[0] == "mf"


def test_plain_filenames_yield_blank_metadata_and_readme_warnings(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "out.xlsx"
    dm = pd.DataFrame(
        {
            "Note": ["C4"],
            "density_metric_raw": [0.5],
            "density_metric_normalized": [1.0],
            "weighted_harmonic_density_contribution": [0.2],
            "weighted_inharmonic_density_contribution": [0.3],
            "weighted_subbass_density_contribution": [0.5],
            "component_harmonic_energy_ratio": [0.2],
            "component_inharmonic_energy_ratio": [0.3],
            "component_subbass_energy_ratio": [0.5],
            "Harmonic Partials sum": [1.0],
            "Inharmonic Partials sum": [2.0],
            "Sub-bass sum": [3.0],
            "Total sum": [6.0],
            "source_file_name": ["recording_plain.wav"],
            "weight_function": ["linear"],
            "density_weighted_sum": [1.5],
            "density_log_weighted": [0.1],
            "effective_partial_density": [0.4],
            "spectral_entropy": [0.7],
        }
    )
    am = pd.DataFrame([("pipeline_contract_version", "test-contract")], columns=["Parameter", "Value"])
    with pd.ExcelWriter(src, engine="openpyxl") as writer:
        dm.to_excel(writer, sheet_name="Density_Metrics", index=False)
        am.to_excel(writer, sheet_name="Analysis_Metadata", index=False)
    proc = _run_export(src, dst)
    assert proc.returncode == 0, proc.stderr + proc.stdout
    sdf = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    if "Instrument" in sdf.columns:
        assert pd.isna(sdf.loc[0, "Instrument"]) or str(sdf.loc[0, "Instrument"]).strip() == ""
    if "Dynamic" in sdf.columns:
        assert pd.isna(sdf.loc[0, "Dynamic"]) or str(sdf.loc[0, "Dynamic"]).strip() == ""
    readme = pd.read_excel(dst, sheet_name="README", engine="openpyxl", header=None)
    text = "\n".join(str(x) for x in readme.iloc[:, 0].tolist() if pd.notna(x))
    assert "Instrument column missing" in text or "could not be inferred" in text
    assert "Dynamic column missing" in text or "could not be inferred" in text


def test_inference_ff_from_filename(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "out.xlsx"
    dm = pd.DataFrame(
        {
            "Note": ["E5"],
            "density_metric_raw": [1.0],
            "density_metric_normalized": [1.0],
            "weighted_harmonic_density_contribution": [0.2],
            "weighted_inharmonic_density_contribution": [0.3],
            "weighted_subbass_density_contribution": [0.5],
            "component_harmonic_energy_ratio": [0.2],
            "component_inharmonic_energy_ratio": [0.3],
            "component_subbass_energy_ratio": [0.5],
            "Harmonic Partials sum": [1.0],
            "Inharmonic Partials sum": [2.0],
            "Sub-bass sum": [3.0],
            "Total sum": [6.0],
            "source_file_name": ["Trumpet_E5_ff.wav"],
            "weight_function": ["linear"],
            "density_weighted_sum": [1.5],
            "density_log_weighted": [0.1],
            "effective_partial_density": [0.4],
            "spectral_entropy": [0.7],
        }
    )
    am = pd.DataFrame([("pipeline_contract_version", "test-contract")], columns=["Parameter", "Value"])
    with pd.ExcelWriter(src, engine="openpyxl") as writer:
        dm.to_excel(writer, sheet_name="Density_Metrics", index=False)
        am.to_excel(writer, sheet_name="Analysis_Metadata", index=False)
    assert _run_export(src, dst).returncode == 0
    df = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    assert df.loc[df["Note"] == "E5", "Instrument"].iloc[0] == "Trumpet"
    assert df.loc[df["Note"] == "E5", "Dynamic"].iloc[0] == "ff"


def test_chart_paths_relative_to_compiled_parent(tmp_path: Path) -> None:
    root = tmp_path
    src = root / "compiled_density_metrics.xlsx"
    _write_minimal_compiled_workbook(src)
    stem_dir = root / "Clarinet_A4_pp" / "A4"
    stem_dir.mkdir(parents=True)
    (stem_dir / "component_amplitude_mass_pie.png").write_bytes(b"\x89PNG\r\n\x1a\n")
    (stem_dir / "component_energy_ratio_pie.png").write_bytes(b"\x89PNG\r\n\x1a\n")
    dst = root / "compiled_density_metrics_research.xlsx"
    assert _run_export(src, dst).returncode == 0
    df = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    row = df.loc[df["Note"] == "A4"].iloc[0]
    amp = str(row["amplitude_mass_chart_file"]).replace("\\", "/")
    erg = str(row["energy_ratio_chart_file"]).replace("\\", "/")
    assert "Clarinet_A4_pp/A4/component_amplitude_mass_pie.png" in amp
    assert "Clarinet_A4_pp/A4/component_energy_ratio_pie.png" in erg


def test_research_export_no_path_columns_canonical_alias_from_v5(tmp_path: Path) -> None:
    """Omit path-like columns; ``canonical_density_v5_adapted`` appears only as ``canonical_density``."""
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "out.xlsx"
    dm = pd.DataFrame(
        {
            "Note": ["A4", "D3"],
            "density_metric_raw": [1.0, 1.0],
            "density_metric_normalized": [1.0, 1.0],
            "weighted_harmonic_density_contribution": [0.2, 0.8],
            "weighted_inharmonic_density_contribution": [0.3, 0.15],
            "weighted_subbass_density_contribution": [0.5, 0.05],
            "component_harmonic_energy_ratio": [0.2, 0.8],
            "component_inharmonic_energy_ratio": [0.3, 0.15],
            "component_subbass_energy_ratio": [0.5, 0.05],
            "Harmonic Partials sum": [1.0, 10.0],
            "Inharmonic Partials sum": [2.0, 20.0],
            "Sub-bass sum": [3.0, 30.0],
            "Total sum": [6.0, 60.0],
            "source_file_name": ["Clarinet_A4_pp.wav", "Bassoon_D3_mf.wav"],
            "weight_function": ["linear", "linear"],
            "density_weighted_sum": [1.5, 2.5],
            "density_log_weighted": [0.1, 0.2],
            "effective_partial_density": [0.4, 0.5],
            "spectral_entropy": [0.7, 0.8],
        }
    )
    canon = pd.DataFrame(
        {
            "Note": ["A4", "D3"],
            "canonical_density_v5_adapted": [1.25, 2.5],
        }
    )
    am = pd.DataFrame(
        [
            ("pipeline_contract_version", "test-contract"),
            ("ANALYSIS_SCHEMA_VERSION", "99"),
            ("weight_function", "linear"),
        ],
        columns=["Parameter", "Value"],
    )
    with pd.ExcelWriter(src, engine="openpyxl") as writer:
        dm.to_excel(writer, sheet_name="Density_Metrics", index=False)
        canon.to_excel(writer, sheet_name="Canonical_Metrics", index=False)
        am.to_excel(writer, sheet_name="Analysis_Metadata", index=False)
    proc = _run_export(src, dst)
    assert proc.returncode == 0, proc.stderr + proc.stdout
    sdm = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    vs = pd.read_excel(dst, sheet_name="Validation_Summary", engine="openpyxl")
    assert "Source_File" not in sdm.columns
    assert "Source_Workbook" not in sdm.columns
    assert "canonical_density_v5_adapted" not in sdm.columns
    assert "canonical_density" not in sdm.columns
    assert "Source_File" not in vs.columns


def test_harmonic_slot_coverage_ratio_matches_slot_matched_over_expected(tmp_path: Path) -> None:
    src = tmp_path / "in.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    assert _run_export(src, dst).returncode == 0
    df = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    needed = {"harmonic_slot_expected_count", "harmonic_slot_matched_count", "harmonic_slot_coverage_ratio"}
    if not needed.issubset(set(df.columns)):
        return
    exp = pd.to_numeric(df["harmonic_slot_expected_count"], errors="coerce")
    det = pd.to_numeric(df["harmonic_slot_matched_count"], errors="coerce")
    ratio = pd.to_numeric(df["harmonic_slot_coverage_ratio"], errors="coerce")
    valid = exp.notna() & det.notna() & ratio.notna() & (exp > 0)
    if valid.any():
        assert np.allclose(ratio[valid], (det[valid] / exp[valid]), equal_nan=True)


def test_energy_families_remain_separate_and_each_sum_to_one(tmp_path: Path) -> None:
    src = tmp_path / "in.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    assert _run_export(src, dst).returncode == 0
    df = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")

    core_cols = {"core_harmonic_energy_ratio", "core_residual_energy_ratio", "core_subbass_energy_ratio"}
    comp_cols = {
        "component_harmonic_energy_ratio",
        "component_inharmonic_energy_ratio",
        "component_subbass_energy_ratio",
    }

    if core_cols.issubset(df.columns):
        core = df[list(core_cols)].apply(pd.to_numeric, errors="coerce")
        valid = core.notna().all(axis=1)
        if valid.any():
            assert np.allclose(core[valid].sum(axis=1), 1.0, atol=1e-6, equal_nan=False)

    if comp_cols.issubset(df.columns):
        comp = df[list(comp_cols)].apply(pd.to_numeric, errors="coerce")
        valid = comp.notna().all(axis=1)
        if valid.any():
            assert np.allclose(comp[valid].sum(axis=1), 1.0, atol=1e-6, equal_nan=False)


def test_body_thickness_columns_and_dashboard_kpis_exist(tmp_path: Path) -> None:
    src = tmp_path / "in.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    assert _run_export(src, dst).returncode == 0
    sdm = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    for c in (
        "body_weighted_effective_density",
        "low_mid_energy_ratio",
        "harmonic_body_density_normalized",
        "residual_body_contribution_capped",
        "spectral_body_thickness_index",
        "salient_harmonic_order_count_up_to_5000hz",
        "expected_harmonic_order_count_up_to_5000hz",
        "salient_harmonic_coverage_up_to_5000hz",
        "salient_harmonic_mass_up_to_5000hz",
        "salient_harmonic_order_count_up_to_density_ceiling_hz",
        "expected_harmonic_order_count_up_to_density_ceiling_hz",
        "salient_harmonic_coverage_up_to_density_ceiling_hz",
        "salient_harmonic_mass_up_to_density_ceiling_hz",
        "salient_odd_harmonic_count_up_to_5000hz",
        "salient_even_harmonic_count_up_to_5000hz",
        "odd_even_harmonic_energy_ratio",
        "salient_inharmonic_log_bin_count_up_to_5000hz",
        "salient_subbass_particle_count",
        "salient_inharmonic_log_bin_count_up_to_density_ceiling_hz",
        "salient_subbass_particle_count_up_to_density_ceiling_hz",
        "final_note_density_count_based",
        "final_note_density_salience_weighted",
        "final_note_density_salience_weighted_norm_for_chart",
        "harmonic_density_component",
        "inharmonic_density_component",
        "subbass_density_component",
        "harmonic_density_weight",
        "inharmonic_density_weight",
        "subbass_density_weight",
        "density_summation_mode",
        "density_salience_threshold_db",
        "density_frequency_ceiling_hz",
    ):
        assert c in sdm.columns
    cd = pd.read_excel(dst, sheet_name="Charts_Data", engine="openpyxl")
    for c in (
        "salient_harmonic_order_count_up_to_5000hz",
        "salient_inharmonic_log_bin_count_up_to_5000hz",
        "salient_subbass_particle_count",
        "final_note_density_count_based",
        "final_note_density_salience_weighted",
        "final_note_density_salience_weighted_norm_for_chart",
        "harmonic_density_component",
        "inharmonic_density_component",
        "subbass_density_component",
    ):
        assert c in cd.columns
    for c in ("spectral_body_thickness_index", "body_weighted_effective_density", "low_mid_energy_ratio"):
        if c in sdm.columns and pd.to_numeric(sdm[c], errors="coerce").notna().any():
            assert c in cd.columns
    if (
        "salient_harmonic_order_count_up_to_5000hz" in sdm.columns
        and pd.to_numeric(sdm["salient_harmonic_order_count_up_to_5000hz"], errors="coerce").notna().any()
    ):
        assert "salient_harmonic_order_count_up_to_5000hz" in cd.columns
    if (
        "salient_inharmonic_log_bin_count_up_to_5000hz" in sdm.columns
        and pd.to_numeric(sdm["salient_inharmonic_log_bin_count_up_to_5000hz"], errors="coerce").notna().any()
    ):
        assert "salient_inharmonic_log_bin_count_up_to_5000hz" in cd.columns
    if (
        "salient_subbass_particle_count" in sdm.columns
        and pd.to_numeric(sdm["salient_subbass_particle_count"], errors="coerce").notna().any()
    ):
        assert "salient_subbass_particle_count" in cd.columns
    if (
        "final_note_density_salience_weighted" in sdm.columns
        and pd.to_numeric(sdm["final_note_density_salience_weighted"], errors="coerce").notna().any()
    ):
        assert "final_note_density_salience_weighted" in cd.columns


def test_metadata_contains_density_and_analysis_controls(tmp_path: Path) -> None:
    src = tmp_path / "in.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    assert _run_export(src, dst).returncode == 0
    md = pd.read_excel(dst, sheet_name="Metadata", engine="openpyxl")
    assert {"Field", "Value"}.issubset(md.columns)
    got = {str(r["Field"]): r["Value"] for _, r in md.iterrows()}
    for k in (
        "density_summation_mode",
        "harmonic_density_weight",
        "inharmonic_density_weight",
        "subbass_density_weight",
        "density_salience_threshold_db",
        "density_frequency_ceiling_hz",
        "window_type",
        "n_fft",
        "hop_length",
        "zero_padding",
        "harmonic_tolerance",
        "frequency_min_hz",
        "frequency_max_hz",
        "magnitude_min_db",
        "source_workbook_sha256",
        "git_commit",
        "git_branch",
        "source_corpus_path",
        "output_path",
    ):
        assert k in got
        assert str(got[k]).strip() != ""


def test_analysis_settings_by_note_sheet_exists_and_is_populated(tmp_path: Path) -> None:
    src = tmp_path / "in.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    assert _run_export(src, dst).returncode == 0
    aset = pd.read_excel(dst, sheet_name="Analysis_Settings_By_Note", engine="openpyxl")
    assert len(aset) == 2
    required = (
        "Note",
        "MIDI",
        "f0_used_for_density_hz",
        "f0_used_for_density_source",
        "acoustic_f0_status",
        "tier_name",
        "n_fft",
        "hop_length",
        "zero_padding",
        "window_type",
        "harmonic_tolerance_hz",
        "frequency_min_hz",
        "frequency_max_hz",
        "magnitude_min_db",
        "magnitude_max_db",
        "density_summation_mode",
        "harmonic_density_weight",
        "inharmonic_density_weight",
        "subbass_density_weight",
        "density_salience_threshold_db",
        "density_frequency_ceiling_hz",
    )
    for c in required:
        assert c in aset.columns
        assert aset[c].astype(str).str.strip().ne("").all()
